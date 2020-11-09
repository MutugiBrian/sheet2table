# IMPORTING THE RELEVANT MODULES
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import mysql.connector

# A COMMON FUNCTION TO BE STORING EXCEL SHEETS TO DB TABLES


def sheet_to_dbtable(cursor, table, sheet, start_row, fund_type, start=1973, finish=2025):

    # INITIAL THE DICTIONARY TO BE TEMPORARILY STORING COUNTRY DATA
    country_data = {}

    # ITERATE THROUGH EACH ROW ON PLEDGES IGNORING THE HEADER
    for key, *values in sheet.iter_rows(min_row=start_row):
        # GET BASIC VALUES
        country_data[key.value] = [v.value for v in values]

        country_iso = key.value
        country_name = values[0].value

        # DELETE THIS COUNTRY'S RECORDED CONTRIBUTION TO AVOID REPETION AND GET FRESH RECORDS
        cursor.execute("DELETE FROM "+table +
                       " WHERE country_iso_2 = '" + country_iso+"'")

        #  GET THIS COUNTRY'S CONTRIBUTION AMOUNT FROM 1973 TO NEXT YEAR
        i = 1
        for year in range(start, finish):
            year_contribution = values[i].value

            if year_contribution is None:
                year_contribution = 0

            if (isinstance(year_contribution, str) and int(year_contribution) <= 0):
                year_contribution = 0

            # STORE THIS COUNTRY AND YEAR CONTRIBUTION TO DB
            val = (country_iso, year, year_contribution,
                   fund_type)
            cursor.execute(
                "INSERT INTO "+table + " (`country_iso_2`, `year`, `amt`, `fund_type`) VALUES (%s, %s, %s, %s);", val)
            i += 1

        # print('________________'+country_name + ' - ' +
        #       country_iso+' DATA INSERTED SUCCESFULLY_______________')
        # print('_________________________________________________________________________________')
        country_data = {}


# xls = pd.ExcelFile(r'contributions.xlsx')
# LOAD EXCEL FILE
xls = load_workbook(filename='contributions.xlsx')


# GET THE PLEDGES AND PAYMENTS
pledges = xls['ef_pledge']
payments = xls['ef_paid']


# GET CURRENT YEAR AND USE IT TO GENERATE NEXT YEAR AS THE MAXIMUM YEAR OF FETCHING CONTRIBUTIONS
current_date = date.today()

next_two_years = current_date.year+2
year_limit = next_two_years

# CONNECT TO THE SQL SERVER
db = mysql.connector.connect(
    host="localhost",
    database="contributions",
    user="root",
    password=""
)

# INITIATE THE DB CONNECTION EXECUTION CURSOR
cursor = db.cursor()


# STORE PLEDGES TO DB USING COMMON FUNCTION
sheet_to_dbtable(cursor, 'pledge',  pledges, 2,
                 'ef', 1973, next_two_years)

# STORE CONTRIBUTIONS TO DB USING COMMON FUNCTION
sheet_to_dbtable(cursor, 'contribution',  payments, 2,
                 'ef', 1973, next_two_years)

# COMMIT ALL CHANGES WHEN DONE
db.commit()

# MESSAGE WHEN DONE
print('EXECUTION SUCCESSFUL - DATA IMPORTED')
